"""
Excel file processing utilities for extracting data from specific cells
"""

import logging
from openpyxl import load_workbook
from typing import Dict, Any, Optional, List

class ExcelProcessor:
    """Class to handle Excel file processing and data extraction"""
    
    def __init__(self):
        self.logger = logging.getLogger(__name__)
    
    def extract_data(self, file_path: str) -> Optional[List[Dict[str, Any]]]:
        """
        Extract data from Excel file all rows starting from row 4, columns A, B, D, E, F, G, I
        
        Args:
            file_path: Path to the Excel file
            
        Returns:
            List of dictionaries with extracted data or None if error
        """
        try:
            # Load workbook
            workbook = load_workbook(file_path, data_only=True)
            
            # Get the first worksheet
            worksheet = workbook.active
            worksheet_name = worksheet.title  # Get the worksheet name
            
            all_rows_data = []
            
            # Process all rows starting from row 4
            for row_num in range(4, worksheet.max_row + 1):
                # Extract data from current row, specific columns
                row_data = {
                    'data': self._get_cell_value(worksheet, f'A{row_num}'),        # Column A - Data
                    'numero_pedido': self._get_cell_value(worksheet, f'B{row_num}'), # Column B - Número do Pedido
                    'nome_cliente': self._get_cell_value(worksheet, f'D{row_num}'),  # Column D - Nome do Cliente
                    'prazo': self._get_cell_value(worksheet, f'E{row_num}'),         # Column E - Prazo
                    'valor_pedido': self._get_cell_value(worksheet, f'F{row_num}'),  # Column F - Valor do Pedido
                    'porcentagem': self._get_cell_value(worksheet, f'G{row_num}'),   # Column G - Porcentagem
                    'frete': self._get_cell_value(worksheet, f'I{row_num}'),         # Column I - Frete
                    'row_number': row_num
                }
                
                # Check if row has significant data (at least valor_pedido or nome_cliente)
                has_data = (
                    (row_data['valor_pedido'] is not None and row_data['valor_pedido'] != 0) or
                    (row_data['nome_cliente'] is not None and str(row_data['nome_cliente']).strip())
                )
                
                if has_data:
                    all_rows_data.append(row_data)
                    self.logger.info(f"Extracted data from row {row_num}: {row_data}")
            
            if not all_rows_data:
                self.logger.warning("No data found in Excel file starting from row 4")
                return None
            
            # Add worksheet name to the data
            result = {
                'worksheet_name': worksheet_name,
                'data': all_rows_data
            }
            
            return result
            
        except Exception as e:
            self.logger.error(f"Error extracting data from Excel file: {str(e)}")
            return None
    
    def _get_cell_value(self, worksheet, cell_address: str) -> Any:
        """
        Get value from a specific cell, handling different data types
        
        Args:
            worksheet: The worksheet object
            cell_address: Cell address (e.g., 'A4')
            
        Returns:
            Cell value or None if empty
        """
        try:
            cell = worksheet[cell_address]
            value = cell.value
            
            # Handle None values
            if value is None:
                return None
            
            # Handle datetime objects - convert to string
            if hasattr(value, 'strftime'):
                return value.strftime('%d/%m/%Y')
            
            # Handle numeric values
            if isinstance(value, (int, float)):
                return value
            
            # Handle string values - strip whitespace
            if isinstance(value, str):
                stripped = value.strip()
                return stripped if stripped else None
            
            # Return as-is for other types
            return value
            
        except Exception as e:
            self.logger.warning(f"Error getting value from cell {cell_address}: {str(e)}")
            return None
    
    def validate_excel_structure(self, file_path: str) -> bool:
        """
        Validate that the Excel file has the expected structure
        
        Args:
            file_path: Path to the Excel file
            
        Returns:
            True if structure is valid, False otherwise
        """
        try:
            workbook = load_workbook(file_path, data_only=True)
            worksheet = workbook.active
            
            # Check if worksheet has at least 4 rows
            if worksheet.max_row < 4:
                self.logger.error("Excel file must have at least 4 rows")
                return False
            
            # Check if worksheet has at least 7 columns (A-G)
            if worksheet.max_column < 7:
                self.logger.error("Excel file must have at least 7 columns (A-G)")
                return False
            
            return True
            
        except Exception as e:
            self.logger.error(f"Error validating Excel structure: {str(e)}")
            return False
    
    def get_preview_data(self, file_path: str, max_rows: int = 10) -> Optional[Dict[str, Any]]:
        """
        Get preview data from Excel file for display purposes
        
        Args:
            file_path: Path to the Excel file
            max_rows: Maximum number of rows to preview
            
        Returns:
            Dictionary with preview data or None if error
        """
        try:
            workbook = load_workbook(file_path, data_only=True)
            worksheet = workbook.active
            
            # Get headers (assuming row 1 contains headers)
            headers = []
            for col in range(1, min(8, worksheet.max_column + 1)):  # Columns A-G
                cell_value = worksheet.cell(row=1, column=col).value
                headers.append(str(cell_value) if cell_value else f"Col {col}")
            
            # Get data rows
            rows = []
            start_row = 2  # Start from row 2 (after headers)
            end_row = min(start_row + max_rows, worksheet.max_row + 1)
            
            for row in range(start_row, end_row):
                row_data = []
                for col in range(1, len(headers) + 1):
                    cell_value = worksheet.cell(row=row, column=col).value
                    row_data.append(str(cell_value) if cell_value else "")
                rows.append(row_data)
            
            return {
                'headers': headers,
                'rows': rows,
                'total_rows': worksheet.max_row,
                'target_row': 4,
                'target_data': self.extract_data(file_path)
            }
            
        except Exception as e:
            self.logger.error(f"Error getting preview data: {str(e)}")
            return None
